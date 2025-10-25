#!/usr/bin/env python3
"""
Import Topogram CSV files from a folder into the local Meteor MongoDB.

Usage:
  ./scripts/import_topograms_folder.py --dir samples/topograms/debian --commit

By default the script runs in dry-run mode and prints counts. Use --commit to write.
It detects Meteor's port from .meteor/local/db/METEOR-PORT or falls back to 3001, or you can
pass an explicit --mongo-url connection string.

This script shells out to `mongosh --port <port> --eval <js>` to perform inserts so it
doesn't require extra Python MongoDB deps and works with the Meteor local DB.

The script understands both the historic Topogram CSV format (node rows beginning with
an id and edge rows with blank first column) and header-based exports that include
`id`, `title`, `source`, `target`, `relationship`, `enlightement`, and similar fields.
It creates a Topogram document per file, then inserts nodes and edges with `topogramId`
referencing the created _id.

Safety: this script will not overwrite existing Topogram documents with the same id.
It generates new ids using ObjectId() in Mongo where needed.
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path


def detect_meteor_port():
    p = Path('.meteor/local/db/METEOR-PORT')
    if p.exists():
        try:
            return int(p.read_text().strip())
        except Exception:
            pass
    return 3001


def find_topogram_files(root):
    rootp = Path(root)
    if not rootp.exists():
        raise SystemExit(f"Folder not found: {root}")
    patterns = ['*.topogram.csv', '*.topogram.xlsx', '*.topogram.ods']
    out = []
    for pat in patterns:
        out.extend([str(p) for p in rootp.rglob(pat)])
    return sorted(out)


def parse_topogram_csv(path):
    nodes = []
    edges = []
    with open(path, newline='', encoding='utf-8') as fh:
        reader = csv.reader(fh)
        header = None
        rows = list(reader)
        if not rows:
            return nodes, edges

        def sanitize(cell):
            if cell is None:
                return ''
            if isinstance(cell, str):
                return cell.strip()
            return str(cell).strip()

        header_map = {}
        if any(isinstance(c, str) and (c.lower() == 'id' or c.lower() == 'title') for c in rows[0]):
            header = [sanitize(c) for c in rows[0]]
            header_map = {}
            for idx, c in enumerate(rows[0]):
                key = sanitize(c)
                if key:
                    header_map[key.lower()] = idx
            body = rows[1:]
        else:
            body = rows

        def get_cell(row, *names, fallback_idx=None):
            if header_map:
                for name in names:
                    if not name:
                        continue
                    idx = header_map.get(name.lower())
                    if idx is not None and idx < len(row):
                        return sanitize(row[idx])
            if fallback_idx is not None and fallback_idx < len(row):
                return sanitize(row[fallback_idx])
            return ''

        for r in body:
            if not any(sanitize(cell) for cell in r):
                continue

            if header_map:
                src = get_cell(r, 'source')
                tgt = get_cell(r, 'target')
                if src or tgt:
                    edge = {'source': src, 'target': tgt, 'raw': r}
                    name = get_cell(r, 'edgelabel', 'edge label')
                    if name:
                        edge['name'] = name
                    color = get_cell(r, 'edgecolor', 'edge color')
                    if color:
                        edge['color'] = color
                    weight = get_cell(r, 'edgeweight', 'edge weight')
                    if weight:
                        edge['weight'] = weight
                    relationship = get_cell(r, 'relationship')
                    if relationship:
                        edge['relationship'] = relationship
                    rel_emoji = get_cell(r, 'relationshipemoji', 'relationship emoji')
                    if rel_emoji:
                        edge['relationshipEmoji'] = rel_emoji
                    enlightement = get_cell(r, 'enlightement', 'enlightenment', 'edgeenlightement', 'edge enlightenment')
                    if enlightement:
                        edge['enlightement'] = enlightement
                    edges.append(edge)
                    continue

            is_edge_row = False
            if len(r) > 0:
                first_cell = sanitize(r[0])
                is_edge_row = (first_cell == '' or first_cell.lower() == 'edge')
            if is_edge_row:
                src = get_cell(r, 'source', fallback_idx=4)
                tgt = get_cell(r, 'target', fallback_idx=5)
                if not src and len(r) > 1:
                    src = get_cell(r, None, fallback_idx=1)
                if not tgt and len(r) > 2:
                    tgt = get_cell(r, None, fallback_idx=2)
                edges.append({'source': src, 'target': tgt, 'raw': r})
                continue

            nid = get_cell(r, 'id', fallback_idx=0)
            title = get_cell(r, 'title', 'name', fallback_idx=1) or nid
            label = get_cell(r, 'label', fallback_idx=2) or title
            node = {'id': nid, 'title': title, 'label': label, 'raw': r}
            emoji = get_cell(r, 'emoji')
            if emoji:
                node['emoji'] = emoji
            color = get_cell(r, 'color', 'fillcolor', 'fill color')
            if color:
                node['color'] = color
            weight = get_cell(r, 'weight', 'rawweight', 'raw weight')
            if weight:
                node['weight'] = weight
            nodes.append(node)
    return nodes, edges


def parse_topogram_spreadsheet(path):
    ext = Path(path).suffix.lower()
    nodes, edges = [], []
    # Try to use openpyxl for .xlsx and pyexcel_ods3 for .ods
    if ext == '.xlsx':
        try:
            import openpyxl  # type: ignore
        except Exception:
            print(f"[WARN] openpyxl not installed; skipping {path} (pip install openpyxl)")
            return nodes, edges
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        lower = [n.lower() for n in names]
        def sheet_to_rows(sh):
            data = []
            it = sh.iter_rows(values_only=True)
            try:
                header = next(it)
            except StopIteration:
                return data
            header = [(str(c).strip() if c is not None else '') for c in header]
            for row in it:
                obj = {}
                for i, h in enumerate(header):
                    if not h:
                        continue
                    val = row[i] if i < len(row) else ''
                    obj[h] = '' if val is None else val
                if any(str(v).strip() for v in obj.values()):
                    data.append(obj)
            return data
        if 'nodes' in lower or 'edges' in lower:
            if 'nodes' in lower:
                sh = wb[names[lower.index('nodes')]]
                nodes = sheet_to_rows(sh)
            if 'edges' in lower:
                sh = wb[names[lower.index('edges')]]
                edges = sheet_to_rows(sh)
        elif names:
            sh = wb[names[0]]
            rows = sheet_to_rows(sh)
            for r in rows:
                keys = [k.lower() for k in r.keys()]
                if 'source' in keys or 'target' in keys or 'from' in keys or 'to' in keys:
                    edges.append(r)
                else:
                    nodes.append(r)
        return nodes, edges
    elif ext == '.ods':
        try:
            from pyexcel_ods3 import get_data  # type: ignore
        except Exception:
            print(f"[WARN] pyexcel-ods3 not installed; skipping {path} (pip install pyexcel-ods3)")
            return nodes, edges
        data = get_data(path)
        names = list(data.keys())
        lower = [n.lower() for n in names]
        def rows_from_array(arr):
            rows = []
            if not arr:
                return rows
            header = [(str(c).strip() if c is not None else '') for c in (arr[0] or [])]
            for r in arr[1:]:
                obj = {}
                for i, h in enumerate(header):
                    if not h:
                        continue
                    val = r[i] if i < len(r) else ''
                    obj[h] = '' if val is None else val
                if any(str(v).strip() for v in obj.values()):
                    rows.append(obj)
            return rows
        if 'nodes' in lower or 'edges' in lower:
            if 'nodes' in lower:
                nodes = rows_from_array(data[names[lower.index('nodes')]])
            if 'edges' in lower:
                edges = rows_from_array(data[names[lower.index('edges')]])
        elif names:
            rows = rows_from_array(data[names[0]])
            for r in rows:
                keys = [k.lower() for k in r.keys()]
                if 'source' in keys or 'target' in keys or 'from' in keys or 'to' in keys:
                    edges.append(r)
                else:
                    nodes.append(r)
        return nodes, edges
    else:
        return nodes, edges


def mongo_insert(target, js_expr, dry_run=True):
    cmd = ['mongosh', '--quiet']
    if target.get('url'):
        cmd.append(target['url'])
    else:
        cmd.extend(['--port', str(target['port'])])
    if dry_run:
        if target.get('url'):
            print('[DRY-RUN] would run mongosh with url', target['url'])
        else:
            print('[DRY-RUN] would run mongosh with port', target['port'])
        return {'ok': True, 'dry': True}

    temp_path = None
    try:
        fd, temp_path = tempfile.mkstemp(prefix='topogram-import-', suffix='.js')
        with os.fdopen(fd, 'w', encoding='utf-8') as handle:
            handle.write(js_expr)
        cmd.extend(['--file', temp_path])
        if target.get('url'):
            print('Running mongosh', target['url'])
        else:
            print('Running mongosh --port', target['port'])
        res = subprocess.run(cmd, check=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

    if res.returncode != 0:
        print('mongosh error:', res.stderr.strip())
        return {'ok': False, 'error': res.stderr.strip(), 'stdout': res.stdout}
    out = res.stdout.strip()
    try:
        return json.loads(out)
    except Exception:
        return {'ok': True, 'raw': out}


def derive_folder_label(dir_path: str) -> str:
    if not dir_path:
        return 'Imported'
    name = Path(dir_path).name.strip()
    if not name:
        return 'Imported'
    sanitized = name.replace('-', ' ').replace('_', ' ')
    words = [w for w in sanitized.split() if w]
    if not words:
        return 'Imported'
    return ' '.join(word.capitalize() for word in words)


def clean_folder(folder_label, mongo_target, dry_run=True):
    folder_json = json.dumps(folder_label)
    js = f"""
(function(){{
  const folderName = {folder_json};
  const tops = db.getCollection('topograms').find({{ folder: folderName }}).toArray();
  if (!tops.length) {{
    print(JSON.stringify({{ ok: true, folder: folderName, deletedTopograms: 0, deletedNodes: 0, deletedEdges: 0 }}));
    return;
  }}
  const objIds = tops.map(t => t._id).filter(id => id instanceof ObjectId);
  const strIds = tops.map(t => {{
    const id = t._id;
    if (!id) return null;
    if (typeof id === 'string') return id;
    if (id instanceof ObjectId) return id.valueOf();
    if (typeof id._str === 'string') return id._str;
    if (typeof id.$oid === 'string') return id.$oid;
    try {{ return String(id); }} catch (e) {{ return null; }}
  }}).filter(Boolean);

  const clauses = [];
  if (objIds.length) {{
    clauses.push({{ topogramId: {{ $in: objIds }} }});
    clauses.push({{ 'data.topogramId': {{ $in: objIds }} }});
  }}
  if (strIds.length) {{
    clauses.push({{ topogramId: {{ $in: strIds }} }});
    clauses.push({{ 'data.topogramId': {{ $in: strIds }} }});
  }}

  let nodesDeleted = 0;
  let edgesDeleted = 0;
  if (clauses.length) {{
    nodesDeleted = db.getCollection('nodes').deleteMany({{ $or: clauses }}).deletedCount;
    edgesDeleted = db.getCollection('edges').deleteMany({{ $or: clauses }}).deletedCount;
  }}
  const topsDeleted = db.getCollection('topograms').deleteMany({{ folder: folderName }}).deletedCount;
  print(JSON.stringify({{ ok: true, folder: folderName, deletedTopograms: topsDeleted, deletedNodes: nodesDeleted, deletedEdges: edgesDeleted }}));
}})();
"""
    return mongo_insert(mongo_target, js, dry_run=dry_run)


def build_and_insert(topogram_name, nodes, edges, mongo_target, folder_label, dry_run=True):
    title = os.path.basename(topogram_name)
    safe_title = json.dumps(title)
    folder_json = json.dumps(folder_label)
    node_payload = []
    for n in nodes:
        node_data = {'id': n['id'], 'title': n['title'], 'label': n['label']}
        for key in ('emoji', 'color', 'weight'):
            if key in n and n[key] not in (None, ''):
                node_data[key] = n[key]
        if 'raw' in n:
            node_data['raw'] = n['raw']
        node_payload.append({'data': node_data})
    edge_payload = []
    for e in edges:
        edge_data = {'source': e['source'], 'target': e['target']}
        for key in ('name', 'label', 'color', 'weight', 'relationship', 'relationshipEmoji', 'enlightement'):
            if key in e and e[key] not in (None, ''):
                edge_data[key] = e[key]
        if 'label' not in edge_data and 'name' in edge_data and edge_data['name']:
            edge_data['label'] = edge_data['name']
        if 'relationship' in edge_data and edge_data['relationship'] and 'label' not in edge_data:
            edge_data['label'] = edge_data['relationship']
        if 'enlightement' not in edge_data or not edge_data['enlightement']:
            edge_data['enlightement'] = 'arrow'
        if 'raw' in e:
            edge_data['raw'] = e['raw']
        edge_payload.append({'data': edge_data})
    nodes_js = json.dumps(node_payload)
    edges_js = json.dumps(edge_payload)
    js = f"""
(function(){{
  const top = {{ title: {safe_title}, source: 'imported-folder', folder: {folder_json}, createdAt: new Date() }};
  top._id = new ObjectId();
  db.getCollection('topograms').insertOne(top);
  const nodelist = {nodes_js};
  nodelist.forEach(n => {{
    if (!n._id) n._id = new ObjectId();
    n.topogramId = top._id;
    n.createdAt = new Date();
    if (!n.data) n.data = {{}};
    n.data.topogramId = top._id;
  }});
  if (nodelist.length) db.getCollection('nodes').insertMany(nodelist);
  const edgelist = {edges_js};
  edgelist.forEach(e => {{
    if (!e._id) e._id = new ObjectId();
    e.topogramId = top._id;
    e.createdAt = new Date();
    if (!e.data) e.data = {{}};
    e.data.topogramId = top._id;
  }});
  if (edgelist.length) db.getCollection('edges').insertMany(edgelist);
    const insertedId = (top._id && typeof top._id.valueOf === 'function') ? top._id.valueOf() : top._id;
    return JSON.stringify({{ ok: true, topogramId: insertedId, nodes: nodelist.length, edges: edgelist.length }});
}})();
"""
    return mongo_insert(mongo_target, js, dry_run=dry_run)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--dir', required=True, help='Folder with .topogram.csv files')
    p.add_argument('--port', type=int, default=None, help='MongoDB port (defaults to Meteor port)')
    p.add_argument('--mongo-url', default=None, help='MongoDB connection string (overrides --port)')
    p.add_argument('--commit', action='store_true', help='Actually write to DB; otherwise dry-run')
    p.add_argument('--limit', type=int, default=None, help='Process only the first N CSV files (testing helper)')
    p.add_argument('--folder', default=None, help='Folder label to assign to imported topograms (defaults to directory name)')
    p.add_argument('--clean-folder', action='store_true', help='Remove existing documents for the folder before import (requires --commit)')
    args = p.parse_args()

    if args.mongo_url and args.port:
        raise SystemExit('Use either --mongo-url or --port, not both')

    if args.mongo_url:
        mongo_target = {'url': args.mongo_url}
    else:
        port = args.port or detect_meteor_port()
        mongo_target = {'port': port}

    files = find_topogram_files(args.dir)
    if not files:
        print('No .topogram.csv files found in', args.dir)
        return

    if args.limit is not None and args.limit >= 0:
        original_count = len(files)
        files = files[:args.limit]
        print(f'Limiting import to first {len(files)} of {original_count} files')

    folder_label = args.folder or derive_folder_label(args.dir)
    print(f'Folder label: "{folder_label}"')

    connection_desc = mongo_target.get('url') or f"port {mongo_target['port']}"
    print(f'Found {len(files)} files in {args.dir}; mongo={connection_desc}; commit={args.commit}')

    if args.clean_folder:
        if not args.commit:
            print('Skipping --clean-folder because --commit was not provided (dry-run).')
        else:
            print(f'Cleaning existing documents for folder "{folder_label}" before import...')
            clean_res = clean_folder(folder_label, mongo_target, dry_run=False)
            print('  cleanup result:', clean_res)

    total_nodes = 0
    total_edges = 0
    for fp in files:
        print('Parsing', fp)
        if fp.lower().endswith('.csv'):
            nodes, edges = parse_topogram_csv(fp)
        else:
            nodes, edges = parse_topogram_spreadsheet(fp)
        print(f'  parsed nodes={len(nodes)}, edges={len(edges)}')
        total_nodes += len(nodes)
        total_edges += len(edges)
        if args.commit:
            res = build_and_insert(fp, nodes, edges, mongo_target, folder_label, dry_run=False)
            print('  insert result:', res)
    print('Summary: files=', len(files), 'nodes=', total_nodes, 'edges=', total_edges)


if __name__ == '__main__':
    main()
